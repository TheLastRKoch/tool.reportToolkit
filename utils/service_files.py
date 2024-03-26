from openpyxl import load_workbook
import pandas
import csv
import os


class ServiceFiles:

    def open_file(self, path):
        os.system("open "+path)

    def read_excel(self, path, sheet_name):
        return pandas.read_excel(path, sheet_name)

    def write_excel(self, path, sheet_name, dataframe):
        dataframe.to_excel(path, sheet_name)

    def clean_excel(self, path, sheet_name):
        wb = load_workbook(path)
        del wb[wb.active.title]
        ws = wb.create_sheet()
        ws.title = sheet_name
        wb.save(path)

    def clean_textfile(self, path):
        os.system("truncate -s 0 "+path)

    def read_textfile(self, path):
        with open(path, "r") as f:
            return f.read()

    def write_textfile(self, path, body):
        with open(path, "a+") as f:
            return f.write(body)

    def json_csv(self, path, json_data):
        with open(path, 'w', newline='') as f:
            f.write("sep=,\n")
            csv_writer = csv.writer(f)

            # Write separator
            header = json_data[0].keys()
            csv_writer.writerow(header)

            for item in json_data:
                csv_writer.writerow(item.values())
